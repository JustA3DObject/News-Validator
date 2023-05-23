import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import re

page_link = []
articles_list = []
title_list = []
link_list = []
category = "world-news"

headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.5060.114 Safari/537.36",
           "X-Amzn-Trace-Id": "Root=1-62d8036d-2b173c1f2e4e7a416cc9e554", "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9",
           "Accept-Encoding": "gzip, deflate, br", "Accept-Language": "en-GB", }

fetch_values_dictionary = {
    "hindustantimes": {
        "domain_link": "https://www.hindustantimes.com/",
        "link": f'https://www.hindustantimes.com/{category}/page-',
        "range_start": 1,
        "range_end": 51,
        "articles_find_element": {
            "element": "section",
            "element_class": "listingPage"
        },
        "articles_findall_element": {
            "element": "div",
            "element_class": "cartHolder",
        },
        "link_find_element": {
            "element": "h3",
            "element_class": "hdg3",
        },
        "content_find_element": {
            "element": "div",
            "element_class": "detailPage",
        },
        "title_find_element": {
            "element": "h1",
            "element_class": "hdg1",
        },
        "datetime_find_element": {
            "element": "div",
            "element_class": "dateTime",
        }
    },
    "ndtv": {
        "link": f'https://www.ndtv.com/{category}/page-',
        "range_start": 1,
        "range_end": 15,
    },
    "theindianexpress": {
        "link": f'https://indianexpress.com/section/{category}/page/',
        "range_start": 2,
        "range_end": 52,
    },
}


def create_newsheet():
    excel = openpyxl.Workbook()
    sheet = excel.active
    sheet.title = 'News'
    sheet.append(['Title', 'Text', 'Subject', 'Date'])


def append_existingsheet():
    excel = openpyxl.load_workbook('True News.xlsx')
    sheet = excel.active


def fetch_hindustantimes_news(fetch_values_dictionary):

    def article_fetcher():

        print("Fetching Articles...")

        for i in range(fetch_values_dictionary["hindustantimes"].get("range_start"),
                       fetch_values_dictionary["hindustantimes"].get("range_end")):
            page_link = (
                f'{fetch_values_dictionary["hindustantimes"].get("link")}{i}')

            i = i+1
            try:
                page = requests.get(page_link, headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                articles = soup2.find(fetch_values_dictionary["hindustantimes"]["articles_find_element"].get("element"),
                                      class_=fetch_values_dictionary["hindustantimes"]["articles_find_element"].get("element_class")).find_all(
                    fetch_values_dictionary["hindustantimes"]["articles_findall_element"].get(
                        "element"),
                    class_=fetch_values_dictionary["hindustantimes"]["articles_findall_element"].get("element_class"))

                articles_list.extend(articles)

            except Exception as e:
                print(e)

    article_fetcher()

    def link_list_maker():

        print("Getting Links...")

        for article in articles_list:

            link = article.find(fetch_values_dictionary["hindustantimes"]["link_find_element"].get("element"),
                                class_=fetch_values_dictionary["hindustantimes"]["link_find_element"].get("element_class"))

            if link is not None:
                link = link.find('a').get('href')

                if (len(re.findall(fetch_values_dictionary["hindustantimes"].get("domain_link"), link)) == 0):
                    link = f'{fetch_values_dictionary["hindustantimes"].get("domain_link")}{link}'
                link_list.append(link)

    link_list_maker()

    def content_fetcher():

        print("Fetching Content...")

        for i in range(len(link_list)):
            try:
                page = requests.get(link_list[i], headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                if fetch_values_dictionary["hindustantimes"]["content_find_element"].get("element_class") is not None:
                    content = soup2.find(
                        fetch_values_dictionary["hindustantimes"]["content_find_element"].get(
                            "element"),
                        class_=fetch_values_dictionary["hindustantimes"]["content_find_element"].get("element_class"))
                elif fetch_values_dictionary["hindustantimes"]["content_find_element"].get("element_id") is not None:
                    content = soup2.find(
                        fetch_values_dictionary["hindustantimes"]["content_find_element"].get(
                            "element"),
                        id=fetch_values_dictionary["hindustantimes"]["content_find_element"].get("element_id"))
                elif fetch_values_dictionary["hindustantimes"]["content_find_element"].get("element_itemprop") is not None:
                    content = soup2.find(
                        fetch_values_dictionary["hindustantimes"]["content_find_element"].get(
                            "element"),
                        itemprop=fetch_values_dictionary["hindustantimes"]["content_find_element"].get("element_itemprop"))

                try:
                    if fetch_values_dictionary["hindustantimes"]["title_find_element"].get("element_class") is not None:
                        title = content.find(
                            fetch_values_dictionary["hindustantimes"]["title_find_element"].get(
                                "element"),
                            class_=fetch_values_dictionary["hindustantimes"]["title_find_element"].get("element_class")).get_text().strip()
                    elif fetch_values_dictionary["hindustantimes"]["title_find_element"].get("element_id") is not None:
                        title = content.find(
                            fetch_values_dictionary["hindustantimes"]["title_find_element"].get(
                                "element"),
                            id=fetch_values_dictionary["hindustantimes"]["title_find_element"].get("element_id")).get_text().strip()
                    elif fetch_values_dictionary["hindustantimes"]["title_find_element"].get("element_id") is not None:
                        title = content.find(
                            fetch_values_dictionary["hindustantimes"]["title_find_element"].get(
                                "element"),
                            itemprop=fetch_values_dictionary["hindustantimes"]["title_find_element"].get("element_itemprop")).get_text().strip()

                    title = ' '.join(title.split())

                except AttributeError:
                    title = "Null"

                try:
                    if fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get("element_class") is not None:
                        date_time = content.find(
                            fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get(
                                "element"), class_=fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get("element_class")).get_text().strip()
                    elif fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get("element_id") is not None:
                        date_time = content.find(
                            fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get(
                                "element"), id=fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get("element_id")).get_text().strip()
                    elif fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get("element_itemprop") is not None:
                        date_time = content.find(
                            fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get(
                                "element"), itemprop=fetch_values_dictionary["hindustantimes"]["datetime_find_element"].get("element_itemprop")).get_text().strip()

                    date_time = date_time.replace('Updated: ', '')

                except AttributeError:
                    date_time = "Null"

                try:
                    body = [x.get_text() for x in content.find_all('p')]
                    body = ' '.join(body)
                    body = ' '.join(body.split())
                    body = body.replace(' Watch Live News: Follow Us:', '')
                    body = body.replace(' ...view detail', '')
                except AttributeError:
                    body = "Null"

                if (title & body & date_time) != "Null":
                    sheet.append([title, body, "Education News", date_time])

            except Exception as e:
                print(e)

    content_fetcher()


def fetch_ndtv_news(fetch_values_dictionary):

    def article_fetcher():

        print("Fetching Articles...")

        for i in range(1, 15):
            page_link = (
                f'{fetch_values_dictionary["ndtv"]["link"]}{i}')

            i = i+1
            try:
                page = requests.get(page_link, headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                articles = soup2.find('div', class_="lisingNews").find_all(
                    'div', class_="news_Itm")

                articles_list.extend(articles)

            except Exception as e:
                print(e)

    article_fetcher()

    def link_list_maker():
        print("Getting Links...")
        for article in articles_list:

            link = article.find('h2', class_='newsHdng')
            if link is not None:
                link = link.find('a').get('href')
                link_list.append(link)

    link_list_maker()

    def content_fetcher():

        print("Fetching Content...")

        for i in range(len(link_list)):
            try:
                page = requests.get(link_list[i], headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                content = soup2.find('section', class_='col-900')

                try:
                    title = content.find(
                        'h1', class_='sp-ttl').get_text().strip()
                    title = ' '.join(title.split())
                except AttributeError:
                    title = "Null"

                try:
                    date_time = content.find(
                        'span', itemprop='dateModified').get_text().strip()
                    date_time = date_time.replace('Updated: ', '')
                except AttributeError:
                    date_time = "Null"

                try:
                    body = [x.get_text() for x in content.find_all('p')]
                    body = ' '.join(body).strip()
                    body = ' '.join(body.split())
                    body = body.replace(' Watch Live News: Follow Us:', '')
                except AttributeError:
                    body = "Null"

                if (title & body & date_time) != "Null":
                    sheet.append([title, body, "Education News", date_time])

            except Exception as e:
                print(e)

    content_fetcher()


def fetch_theindianexpress_news(fetch_values_dictionary):

    def article_fetcher():

        print("Fetching Articles...")

        for i in range(2, 52):
            page_link = (
                f'{fetch_values_dictionary["theindianexpress"]["link"]}{i}')

            i = i+1
            try:
                page = requests.get(page_link, headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                articles = soup2.find('div', class_='nation').find_all(
                    'div', class_='articles')

                articles_list.extend(articles)

            except Exception as e:
                print(e)

    article_fetcher()

    def link_list_maker():

        print("Getting Links...")

        for article in articles_list:

            link = article.find('h2', class_='title')
            if link is not None:
                link = link.find('a').get('href')
                link_list.append(link)

    link_list_maker()

    def content_fetcher():

        print("Fetching Content...")

        for i in range(len(link_list)):
            try:
                page = requests.get(link_list[i], headers=headers)
                page.raise_for_status()

                if page.status_code == 200:
                    soup1 = BeautifulSoup(page.text, "html.parser")
                    soup2 = BeautifulSoup(soup1.prettify(), "html.parser")

                content = soup2.find('div', class_='container native_story')

                try:
                    title = content.find(
                        'h1', itemprop='headline').get_text().strip()
                    title = ' '.join(title.split())
                except AttributeError:
                    title = "Null"

                try:
                    date_time = content.find(
                        'span', itemprop='dateModified').get_text().strip()
                    date_time = date_time.replace('Updated: ', '')
                except AttributeError:
                    date_time = "Null"

                try:
                    body = [x.get_text() for x in content.find_all('p')]
                    body = ' '.join(body).strip()
                    body = ' '.join(body.split())
                except AttributeError:
                    body = "Null"

                if (title & body & date_time) != "Null":
                    sheet.append([title, body, "Education News", date_time])

            except Exception as e:
                print(e)

    content_fetcher()


# create_newsheet()

# append_existingsheet()

# fetch_hindustantimes_news(fetch_values_dictionary)

# fetch_ndtv_news(fetch_values_dictionary)

# fetch_theindianexpress_news(fetch_values_dictionary)

# print("Saving Excel File...")
# excel.save('True News.xlsx')
